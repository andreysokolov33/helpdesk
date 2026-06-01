from __future__ import annotations

import asyncio
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from fastapi import HTTPException, Request
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.routers.helpdesk.operator_log_service import write_operator_log
from app.api.v1.routers.helpdesk.traffic_detail_schemas import (
    TrafficDetailSendRequest,
    TrafficDetailSendResponse,
)
from app.api.v1.routers.helpdesk.user_profile_service import _load_personal_with_balance
from app.config import settings
from app.core.smtp_skystream import send_skystream_email, smtp_skystream_is_configured

_TEMPLATES_DIR = Path(settings.TEMPLATES_DIR)
_jinja = Environment(
    loader=FileSystemLoader(str(_TEMPLATES_DIR / "emails")),
    autoescape=select_autoescape(enabled_extensions=("html", "xml")),
)

_ACTIVE_SESSION_LABEL = "Активная сессий"
def _pick_recipient_email(email: Optional[str], is_juridical: int) -> Optional[str]:
    if not email or not str(email).strip():
        return None
    raw = str(email).strip()
    if is_juridical == 2:
        parts = [p.strip() for p in raw.split(";") if p.strip()]
        return parts[0] if parts else None
    return raw


def _fmt_date_ru(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def _fmt_dt_ru(dt: Optional[datetime]) -> str:
    if dt is None:
        return _ACTIVE_SESSION_LABEL
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt.strftime("%d.%m.%Y %H:%M:%S")


def _safe_filename_part(value: str) -> str:
    return re.sub(r"[^\w\-.]+", "_", value, flags=re.UNICODE)[:64]


async def _fetch_sessions(
    session: AsyncSession,
    login: str,
    date_from: date,
    date_to: date,
) -> list[dict[str, Any]]:
    end_exclusive = date_to + timedelta(days=1)
    rows = (
        await session.execute(
            text("""
                SELECT
                    acctstarttime,
                    acctstoptime,
                    acctinputoctets / 1024.0 / 1024.0 AS in_mb,
                    acctoutputoctets / 1024.0 / 1024.0 AS out_mb,
                    (acctinputoctets + acctoutputoctets) / 1024.0 / 1024.0 AS total_mb,
                    framedipaddress::text AS framedipaddress,
                    CASE WHEN framedprotocol = 'PPP' THEN 'PPPoE' ELSE 'HOTSPOT' END AS protocol
                FROM radius.radacct r
                WHERE lower(r.username) = lower(:login)
                  AND acctstarttime >= :date_from
                  AND acctstarttime < :date_to_exclusive
                ORDER BY acctstarttime
            """),
            {
                "login": login.strip(),
                "date_from": datetime.combine(date_from, datetime.min.time()),
                "date_to_exclusive": datetime.combine(end_exclusive, datetime.min.time()),
            },
        )
    ).mappings().all()
    return [dict(r) for r in rows]


async def _fetch_daily(
    session: AsyncSession,
    user_id: int,
    date_from: date,
    date_to: date,
) -> list[dict[str, Any]]:
    rows = (
        await session.execute(
            text("""
                SELECT
                    rhed."date",
                    rhed.download,
                    rhed.upload,
                    rhed.mb,
                    CASE WHEN rhed.protocol = 'PPP' THEN 'PPPoE' ELSE 'HOTSPOT' END AS protocol
                FROM traffic.radacct_history_every_day rhed
                WHERE rhed.user_id = :uid
                  AND rhed."date" >= :date_from
                  AND rhed."date" <= :date_to
                ORDER BY rhed."date"
            """),
            {"uid": user_id, "date_from": date_from, "date_to": date_to},
        )
    ).mappings().all()
    return [dict(r) for r in rows]


async def _fetch_hourly(
    session: AsyncSession,
    user_id: int,
    date_from: date,
    date_to: date,
) -> list[dict[str, Any]]:
    rows = (
        await session.execute(
            text("""
                SELECT
                    rheh."date",
                    rheh."hour",
                    rheh.download,
                    rheh.upload,
                    rheh.mb,
                    CASE WHEN rheh.protocol = 'PPP' THEN 'PPPoE' ELSE 'HOTSPOT' END AS protocol
                FROM traffic.radacct_history_every_hour rheh
                WHERE rheh.user_id = :uid
                  AND rheh."date" >= :date_from
                  AND rheh."date" <= :date_to
                ORDER BY rheh."date", rheh."hour"
            """),
            {"uid": user_id, "date_from": date_from, "date_to": date_to},
        )
    ).mappings().all()
    return [dict(r) for r in rows]


def _autosize_columns(ws, min_width: int = 10, max_width: int = 48) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _write_sheet_with_totals(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    numeric_cols: set[int],
) -> None:
    bold = Font(bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    if not rows:
        _autosize_columns(ws)
        return

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="Итого").font = bold
    first_data = 2
    last_data = len(rows) + 1
    for c_idx in range(1, len(headers) + 1):
        if c_idx not in numeric_cols:
            continue
        letter = get_column_letter(c_idx)
        formula = f"=SUM({letter}{first_data}:{letter}{last_data})"
        cell = ws.cell(row=total_row, column=c_idx, value=formula)
        cell.font = bold
        if c_idx in numeric_cols:
            cell.number_format = "#,##0.00"

    _autosize_columns(ws)


def _build_excel(
    *,
    sessions: list[dict[str, Any]],
    daily: list[dict[str, Any]],
    hourly: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Детализация сессий")
    session_rows: list[list[Any]] = []
    for r in sessions:
        stop = r.get("acctstoptime")
        session_rows.append(
            [
                _fmt_dt_ru(r.get("acctstarttime")),
                _ACTIVE_SESSION_LABEL if stop is None else _fmt_dt_ru(stop),
                round(float(r.get("in_mb") or 0), 4),
                round(float(r.get("out_mb") or 0), 4),
                round(float(r.get("total_mb") or 0), 4),
                r.get("framedipaddress") or "",
                r.get("protocol") or "",
            ]
        )
    _write_sheet_with_totals(
        ws1,
        [
            "Начало сессии",
            "Окончание сессии",
            "Входящий, МБ",
            "Исходящий, МБ",
            "Всего, МБ",
            "IP-адрес",
            "Протокол",
        ],
        session_rows,
        numeric_cols={3, 4, 5},
    )

    ws2 = wb.create_sheet("Суточный трафик")
    daily_rows: list[list[Any]] = []
    for r in daily:
        d = r.get("date")
        daily_rows.append(
            [
                d.strftime("%d.%m.%Y") if isinstance(d, date) else str(d or ""),
                round(float(r.get("download") or 0), 2),
                round(float(r.get("upload") or 0), 2),
                round(float(r.get("mb") or 0), 2),
                r.get("protocol") or "",
            ]
        )
    _write_sheet_with_totals(
        ws2,
        ["Дата", "Скачано", "Отдано", "Всего, МБ", "Протокол"],
        daily_rows,
        numeric_cols={2, 3, 4},
    )

    ws3 = wb.create_sheet("Почасовой трафик")
    hourly_rows: list[list[Any]] = []
    for r in hourly:
        d = r.get("date")
        hourly_rows.append(
            [
                d.strftime("%d.%m.%Y") if isinstance(d, date) else str(d or ""),
                int(r["hour"]) if r.get("hour") is not None else "",
                round(float(r.get("download") or 0), 2),
                round(float(r.get("upload") or 0), 2),
                round(float(r.get("mb") or 0), 2),
                r.get("protocol") or "",
            ]
        )
    _write_sheet_with_totals(
        ws3,
        ["Дата", "Час (МСК)", "Скачано", "Отдано", "Всего, МБ", "Протокол"],
        hourly_rows,
        numeric_cols={3, 4, 5},
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_email_html(
    *,
    user_id: int,
    period_from: str,
    period_to: str,
) -> str:
    tpl = _jinja.get_template("traffic_detail.html")
    return tpl.render(
        user_id=user_id,
        period_from=period_from,
        period_to=period_to,
    )


def _email_plain(
    *,
    user_id: int,
    period_from: str,
    period_to: str,
) -> str:
    return (
        "Здравствуйте!\n\n"
        f"Во вложении — детализация расхода трафика по вашей учётной записи "
        f"(ID: {user_id}) за период с {period_from} по {period_to}.\n\n"
        "Время в файле указано по московскому часовому поясу (МСК, UTC+3).\n\n"
        "Содержимое файла:\n"
        "1. Лист «Детализация сессий» — список сессий подключения с объёмом трафика и IP-адресом.\n"
        "2. Лист «Суточный трафик» — суммарный расход трафика за каждый день периода.\n"
        "3. Лист «Почасовой трафик» — расход трафика по часам (МСК) за выбранные даты.\n\n"
        "Внизу каждого листа приведены итоговые суммы по числовым столбцам.\n\n"
        "Это автоматическое письмо, отвечать на него не нужно.\n"
        "Служба поддержки WiFiТочка"
    )


async def send_traffic_detail(
    session: AsyncSession,
    user_id: int,
    body: TrafficDetailSendRequest,
    operator: dict[str, Any],
    request: Request,
) -> TrafficDetailSendResponse:
    if not smtp_skystream_is_configured():
        raise HTTPException(
            status_code=503,
            detail="Отправка писем не настроена (SMTP SkyStream)",
        )

    date_from = body.date_from
    date_to = body.date_to
    if date_to < date_from:
        raise HTTPException(status_code=400, detail="Дата окончания раньше даты начала")

    personal, _balance = await _load_personal_with_balance(session, user_id)
    recipient = _pick_recipient_email(personal.email, personal.is_juridical)
    if not recipient:
        raise HTTPException(
            status_code=400,
            detail="В профиле абонента не указан e-mail для отправки детализации",
        )

    login = (personal.login or "").strip()
    if not login:
        raise HTTPException(status_code=400, detail="У абонента не задан логин")

    sessions, daily, hourly = await asyncio.gather(
        _fetch_sessions(session, login, date_from, date_to),
        _fetch_daily(session, user_id, date_from, date_to),
        _fetch_hourly(session, user_id, date_from, date_to),
    )

    xlsx_bytes = _build_excel(sessions=sessions, daily=daily, hourly=hourly)
    period_from = _fmt_date_ru(date_from)
    period_to = _fmt_date_ru(date_to)
    subject = f"Детализация трафика WiFiТочка за {period_from} — {period_to}"
    fname = (
        f"traffic_detail_{_safe_filename_part(login)}_"
        f"{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    )

    html = _render_email_html(
        user_id=user_id,
        period_from=period_from,
        period_to=period_to,
    )
    plain = _email_plain(
        user_id=user_id,
        period_from=period_from,
        period_to=period_to,
    )

    try:
        await asyncio.to_thread(
            send_skystream_email,
            to_addr=recipient,
            subject=subject,
            body_text=plain,
            body_html=html,
            attachments=[
                (
                    xlsx_bytes,
                    "application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    fname,
                )
            ],
        )
    except Exception as exc:
        await write_operator_log(
            session,
            operator_id=int(operator["user_id"]),
            action="traffic_detail.send",
            subscriber_id=user_id,
            page=f"/users/{user_id}",
            request=request,
            details={
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
                "email": recipient,
            },
            success=False,
            error_message=str(exc),
            auto_commit=False,
        )
        await session.commit()
        raise HTTPException(
            status_code=502,
            detail=f"Не удалось отправить письмо: {exc}",
        ) from exc

    await write_operator_log(
        session,
        operator_id=int(operator["user_id"]),
        action="traffic_detail.send",
        subscriber_id=user_id,
        page=f"/users/{user_id}",
        request=request,
        details={
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "email": recipient,
            "sessions_count": len(sessions),
            "daily_count": len(daily),
            "hourly_count": len(hourly),
        },
        auto_commit=False,
    )
    await session.commit()

    return TrafficDetailSendResponse(
        message="Детализация отправлена на e-mail абонента",
        email=recipient,
    )
